import os
import sys
import requests
import bs4
from bs4 import BeautifulSoup
import openpyxl


def downloadData(search_term):
    job_titles = list()
    locations = list()
    companies = list()
    summary = list()
    salaries = list()

    cur_dir = os.getcwd()  # Save current directory
    os.makedirs(cur_dir + '/downloads/data/', exist_ok=True)  # Make the directory if it doesn't exist
    stats_file_path = cur_dir + '/downloads/data/indeed_data.xlsx'  # Store path
    wb = openpyxl.Workbook()  # Create a new workbook object
    wb.save(stats_file_path)  # Save the file

    for i in range(0,100,10):
        url = 'https://www.indeed.com/jobs?q=' + search_term + "&start=" + str(i)
        response = requests.get(url)
        data = response.text  # Convert to text
        soup = BeautifulSoup(data, 'html.parser')  # Parse the html response

        divs = soup.find_all(name="div", attrs={"class": "row"})

        #Get the job titles
        for div in divs:
            for a in div.find_all(name="a", attrs={"data-tn-element": "jobTitle"}):
                job_titles.append(a["title"])

        #Get the locations
        for a in soup.find_all(name="span", attrs={"class": "location"}):
            locations.append(a.text)

        #Get the company names
        for div in divs:
            for a in div.find_all(name="span", attrs={"class": "company"}):
                companies.append(a.text.strip())

        #Get job summary
        for a in soup.find_all(name="span", attrs={"class": "summary"}):
            summary.append(a.text.strip())

        #Get salary
        for div in soup.find_all(name="div", attrs={"class": "row"}):
            try:
                div_two = div.find(name="div", attrs={"class": "sjcl"})
                div_three = div_two.find("div")
                salaries.append(div_three.text.strip())
            except:
                salaries.append("Not available")

    wb = openpyxl.load_workbook(stats_file_path)
    wb.create_sheet(title=search_term)
    sheet = wb[search_term]
    job_titles = ['Job Title'] + job_titles
    companies = ['Company Name'] + companies
    locations = ['Location'] + locations
    summary = ['Summary'] + summary
    salaries = ['Salary'] + salaries
    for i in range(1, len(job_titles)):
        sheet.cell(row = i, column = 1).value = job_titles[i-1]
        sheet.cell(row=i, column=2).value = companies[i - 1]
        sheet.cell(row=i, column=3).value = locations[i - 1]
        sheet.cell(row=i, column=4).value = summary[i - 1]
        sheet.cell(row=i, column=5).value = salaries[i - 1]

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
        
    wb.save(stats_file_path)

if __name__ == "__main__":
    os.chdir(os.path.dirname(sys.argv[0]))
    search_term = input("Enter search term:")
    downloadData(search_term)
